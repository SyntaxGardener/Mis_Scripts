# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import os
import threading
import platform
import subprocess

class ProcesadorNotasESPAD:
    def __init__(self):
        self.mapeo_materias = {
            '1.2': {'LED': 'LEN', 'AUL': 'LIT', 'USLE': 'ING-1', 'ASLE': 'ING-2'},
            '2.1': {'ENL': 'LEN', 'PAL': 'LIT', 'LELE': 'ING-1', 'SOLE': 'ING-2'},
            '2.2': {'CPL': 'LEN', 'LIM': 'LIT', 'ECLE': 'ING-1', 'INLE': 'ING-2'}
        }
        self.config_notas = {
            'LEN': [('Examen', 6), ('Auto', 0.5), ('Oral', 0.5), ('Escrita', 1), ('Invest', 2)],
            'LIT': [('Examen', 6), ('Auto', 1), ('Escrita', 1.5), ('Lectura', 1.5)],
            'ING-1': [('Examen', 6), ('Auto', 0.5), ('Oral', 1.5), ('Ejerc', 2)],
            'ING-2': [('Examen', 6), ('Auto', 0.5), ('Oral', 1.5), ('Ejerc', 2)]
        }
        self.niveles = ['1.2', '2.1', '2.2']

    def extraer(self, ruta):
        alumnos = {n: [] for n in self.niveles}
        registro_bloqueos = {} 
        with pdfplumber.open(ruta) as pdf:
            for page in pdf.pages:
                texto = page.extract_text() or ""
                nivel_actual = next((n for n in self.niveles if f"Nivel {n}" in texto), None)
                if not nivel_actual: continue
                tablas = page.extract_tables()
                for tabla in tablas:
                    if not tabla or "Alumnado" not in str(tabla[0]): continue
                    encabezado = [str(c).strip().upper() for c in tabla[0]]
                    for fila in tabla[1:]:
                        if not fila or not fila[0] or "TOTAL:" in str(fila[0]).upper(): continue
                        nombre_raw = str(fila[0]).strip()
                        limpio = re.sub(r'ESPAD Mixt|ESP\.-SA|MATR|APRO|EXEN|\d{5,}', '', nombre_raw, flags=re.I).strip().strip('-').strip()
                        if ',' in limpio:
                            ape, nom = [part.strip() for part in limpio.split(',', 1)]
                            ape = ape.upper()
                            alumnos[nivel_actual].append((ape, nom))
                            mats_a_bloquear = []
                            for i, celda in enumerate(fila):
                                valor = str(celda).strip().upper() if celda else ""
                                if valor in ["APRO", "EXEN", ""]:
                                    cod_pdf = encabezado[i]
                                    if cod_pdf in self.mapeo_materias[nivel_actual]:
                                        mats_a_bloquear.append(self.mapeo_materias[nivel_actual][cod_pdf])
                            registro_bloqueos[(nivel_actual, ape, nom)] = mats_a_bloquear
        return alumnos, registro_bloqueos

    def generar_excel(self, datos, bloqueos, ruta):
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            for n in self.niveles:
                hoja = f'Nivel_{n}'.replace('.', '_')
                cols = ['Apellidos', 'Nombre']
                for asig in ['LEN', 'LIT', 'ING-1', 'ING-2']:
                    for c_nom, peso in self.config_notas[asig]: 
                        cols.append(f"{asig}_{c_nom} ({peso})")
                    cols.append(f"{asig}_TOT")
                df = pd.DataFrame(sorted(list(set(datos[n]))), columns=['Apellidos', 'Nombre'])
                for c in cols[2:]: df[c] = ""
                df.to_excel(writer, sheet_name=hoja, index=False)
        self.estilizar(ruta, bloqueos)

    def estilizar(self, ruta, bloqueos):
        wb = load_workbook(ruta)
        gris_bloqueo = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        azul_h = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        azul_f = PatternFill(start_color="E7EBF5", end_color="E7EBF5", fill_type="solid")
        f_blanca = Font(color="FFFFFF", bold=True)
        borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for ws in wb.worksheets:
            nivel_hoja = ws.title.replace('Nivel_', '').replace('_', '.')
            for cell in ws[1]:
                cell.fill, cell.font, cell.alignment = azul_h, f_blanca, Alignment(horizontal='center', vertical='center')
            for r_idx in range(2, ws.max_row + 1):
                ape, nom = str(ws.cell(row=r_idx, column=1).value), str(ws.cell(row=r_idx, column=2).value)
                mats_gris = bloqueos.get((nivel_hoja, ape, nom), [])
                for c_idx in range(3, ws.max_column + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.border = borde
                    header = str(ws.cell(row=1, column=c_idx).value)
                    asig_actual = header.split('_')[0]
                    if asig_actual in mats_gris:
                        cell.fill = gris_bloqueo
                    else:
                        if header.endswith("_TOT"):
                            cell.fill = azul_f
                            prefix = header.split('_')[0]
                            start_c = next(i for i in range(3, c_idx) if str(ws.cell(row=1, column=i).value).startswith(prefix))
                            c1, c2 = ws.cell(row=r_idx, column=start_c).column_letter, ws.cell(row=r_idx, column=c_idx-1).column_letter
                            cell.value = f"=SUM({c1}{r_idx}:{c2}{r_idx})"
                    cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 25 if col[0].column <= 2 else 15
        wb.save(ruta)

class AppAsistenteFinal:
    def __init__(self, root):
        self.root = root
        self.root.title("Generador de Registro de Notas (excel)")
        ancho = 650
        alto = 480
        pos_x = (self.root.winfo_screenwidth() // 2) - (ancho // 2)
        self.root.geometry(f"{ancho}x{alto}+{pos_x}+0")
        
        self.archivo_pdf = ""
        self.carpeta_destino = ""
        self.nombre_excel = ""
        self.abrir_al_final = tk.BooleanVar(value=True) # Activado por defecto

        tk.Label(root, text="GENERADOR DE REGISTRO DE NOTAS", font=("Arial", 13, "bold"), fg="#1A237E").pack(pady=20)

        # PASO 1: Elegir PDF
        self.frame1 = tk.LabelFrame(root, text=" Paso 1: Seleccionar PDF (alumnos y materias)", padx=10, pady=10)
        self.frame1.pack(pady=10, padx=30, fill="x")
        self.lbl_pdf = tk.Label(self.frame1, text="Esperando archivo...", fg="#757575", font=("Arial", 9, "italic"))
        self.lbl_pdf.pack(side="left", padx=5)
        self.btn_pdf = tk.Button(self.frame1, text="Elegir PDF", command=self.seleccionar_pdf, bg="#E1E2E1", width=12)
        self.btn_pdf.pack(side="right")

        # PASO 2: Donde y Como
        self.frame2 = tk.LabelFrame(root, text=" Paso 2: Destino y nombre del archivo ", padx=10, pady=10)
        self.frame2.pack(pady=10, padx=30, fill="x")
        self.lbl_dest = tk.Label(self.frame2, text="Complete el paso 1 primero", fg="#BDBDBD", font=("Arial", 8))
        self.lbl_dest.pack(side="top", anchor="w", padx=5)
        
        self.btn_dest = tk.Button(self.frame2, text="Elegir Carpeta", command=self.seleccionar_destino, bg="#E1E2E1", width=12)
        self.btn_dest.pack(side="right", pady=5)

        # PASO 3: Ejecutar y Tick
        self.frame3 = tk.Frame(root)
        self.frame3.pack(pady=20)
        
        self.chk_abrir = tk.Checkbutton(self.frame3, text="Abrir carpeta automaticamente al terminar", 
                                       variable=self.abrir_al_final, font=("Arial", 9))
        self.chk_abrir.pack()

        self.btn_procesar = tk.Button(root, text="PASO 3: PROCESAR TODO", command=self.iniciar_proceso, 
                                     bg="#E1E2E1", fg="black", font=("Arial", 11, "bold"), 
                                     height=2, width=30)
        self.btn_procesar.pack(pady=10)

        self.lbl_status = tk.Label(root, text="", font=("Arial", 9, "bold"))
        self.lbl_status.pack()

    def seleccionar_pdf(self):
        archivo = filedialog.askopenfilename(title="Selecciona el PDF", filetypes=[("PDF", "*.pdf")])
        if archivo:
            self.archivo_pdf = archivo
            # Extraer el nombre sin extension para el excel
            nombre_base = os.path.splitext(os.path.basename(archivo))[0]
            self.nombre_excel = f"{nombre_base}.xlsx"
            
            self.lbl_pdf.config(text=os.path.basename(archivo), fg="#2E7D32", font=("Arial", 9, "bold"))
            self.btn_dest.config(bg="#3F51B5", fg="white", cursor="hand2")
            self.lbl_dest.config(text=f"Generara: {self.nombre_excel}", fg="black", font=("Arial", 9, "italic"))

    def seleccionar_destino(self):
        if not self.archivo_pdf: return
        directorio = filedialog.askdirectory()
        if directorio:
            self.carpeta_destino = directorio
            self.lbl_dest.config(text=f"En: {directorio}\nNombre: {self.nombre_excel}", fg="#2E7D32", font=("Arial", 8, "bold"))
            self.btn_procesar.config(bg="#2E7D32", fg="white", cursor="hand2")

    def iniciar_proceso(self):
        if not self.archivo_pdf or not self.carpeta_destino:
            return
        self.btn_procesar.config(text="TRABAJANDO...", bg="#FFA000")
        self.lbl_status.config(text="Extrayendo datos y aplicando estilos...", fg="#1A237E")
        threading.Thread(target=self.ejecutar, daemon=True).start()

    def ejecutar(self):
        try:
            p = ProcesadorNotasESPAD()
            datos, bloqueos = p.extraer(self.archivo_pdf)
            ruta_completa = os.path.join(self.carpeta_destino, self.nombre_excel)
            p.generar_excel(datos, bloqueos, ruta_completa)
            
            self.root.after(0, self.finalizar_exito)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.root.after(0, lambda: self.btn_procesar.config(text="REINTENTAR", bg="#D32F2F"))

    def finalizar_exito(self):
        self.btn_procesar.config(text="¡COMPLETADO!", bg="#2E7D32")
        self.lbl_status.config(text="Archivo generado con exito", fg="#2E7D32")
        
        if self.abrir_al_final.get():
            self.abrir_explorador()
        else:
            messagebox.showinfo("Hecho", f"El archivo {self.nombre_excel} esta listo.")

    def abrir_explorador(self):
        if platform.system() == "Windows": os.startfile(self.carpeta_destino)
        elif platform.system() == "Darwin": subprocess.Popen(["open", self.carpeta_destino])
        else: subprocess.Popen(["xdg-open", self.carpeta_destino])

if __name__ == "__main__":
    root = tk.Tk()
    app = AppAsistenteFinal(root)
    root.mainloop()