import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import subprocess
import sys
from datetime import datetime

# ── Dependencias opcionales ──────────────────────────────────────────────────
try:
    import pandas as pd
except ImportError:
    import subprocess as _sp
    _sp.run([sys.executable, "-m", "pip", "install", "pandas", "openpyxl", "xlrd"], check=True)
    import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess as _sp
    _sp.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

# ── Paleta de colores ────────────────────────────────────────────────────────
C_BG        = "#F7F9FC"
C_PANEL     = "#FFFFFF"
C_ACCENT    = "#2563EB"
C_SUCCESS   = "#16A34A"
C_TEXT      = "#1E293B"
C_MUTED     = "#64748B"
C_BORDER    = "#E2E8F0"
C_BTN_FG    = "#FFFFFF"

# ── Lógica de datos ──────────────────────────────────────────────────────────

NAT_MAP = {
    "Alemana": "Alemania", "Argentina": "Argentina", "Belga": "Bélgica",
    "Bielorrusa": "Bielorrusia", "Brasileña": "Brasil", "Colombiana": "Colombia",
    "Cubana": "Cuba", "Estadounidense": "EEUU", "Gambiana": "Gambia",
    "Guatemalteca": "Guatemala", "Marroquí": "Marruecos", "Nicaragüense": "Nicaragua",
    "Nigeriana": "Nigeria", "Paraguaya": "Paraguay", "Peruana": "Perú",
    "Portuguesa": "Portugal", "Rumana": "Rumanía", "Rusa": "Rusia",
    "Ucraniana": "Ucrania", "Venezolana": "Venezuela",
    # variantes adicionales
    "Alemán": "Alemania", "Argentino": "Argentina", "Brasileño": "Brasil",
    "Colombiano": "Colombia", "Cubano": "Cuba", "Gambiano": "Gambia",
    "Guatemalteco": "Guatemala", "Marroquí": "Marruecos",
    "Nigeriano": "Nigeria", "Paraguayo": "Paraguay", "Peruano": "Perú",
    "Portugués": "Portugal", "Rumano": "Rumanía",
    "Ucraniano": "Ucrania", "Venezolano": "Venezuela",
}

AGE_LABELS = [
    "16 años", "17 años", "18 años", "19 años",
    "20 años", "21 años", "22 años", "23 años",
    "24 años", "25 años", "26 años", "27 años",
    "28 años", "29 años",
    "30-34 años", "35-39 años", "40-49 años", "50-64 años", "≥ 65 años",
]

def _age_bucket(age):
    if age is None:
        return None
    a = int(age)
    if a < 16:
        return "< 16 años"
    if a <= 29:
        return f"{a} años"
    if a <= 34: return "30-34 años"
    if a <= 39: return "35-39 años"
    if a <= 49: return "40-49 años"
    if a <= 64: return "50-64 años"
    return "≥ 65 años"

def _parse_date(val):
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

def procesar_listado(path, fecha_ref=None):
    if fecha_ref is None:
        fecha_ref = datetime(datetime.now().year, 9, 1)
    ext = os.path.splitext(path)[1].lower()

    # Columnas que buscamos (en mayúsculas)
    COLS_CLAVE = {"SEXO", "GÉNERO", "GENERO", "NACIMIENTO", "F.NAC", "FECHA_NAC",
                  "FECHA NAC", "FECHA DE NACIMIENTO", "F_NAC", "BIRTHDATE", "DOB",
                  "NOMBRE", "ALUMNO", "APELLIDOS", "NACIONALIDAD", "NACION"}

    def _leer(header_row=0):
        if ext == ".xls":
            return pd.read_excel(path, engine="xlrd", header=header_row)
        elif ext in (".xlsx", ".xlsm"):
            return pd.read_excel(path, engine="openpyxl", header=header_row)
        elif ext == ".csv":
            for sep in [";", ",", "\t"]:
                try:
                    df = pd.read_csv(path, sep=sep, header=header_row)
                    if len(df.columns) > 2:
                        return df
                except Exception:
                    pass
        raise ValueError(f"Formato no soportado: {ext}")

    # Intentar primero con header=0; si no reconoce columnas clave,
    # buscar en las primeras 10 filas cuál contiene las cabeceras reales
    df = _leer(0)
    cols_upper = {str(c).strip().upper() for c in df.columns}
    if not cols_upper & COLS_CLAVE:
        encontrado = False
        for fila in range(1, 10):
            df_prueba = _leer(fila)
            cols_prueba = {str(c).strip().upper() for c in df_prueba.columns}
            if cols_prueba & COLS_CLAVE:
                df = df_prueba
                encontrado = True
                break
        if not encontrado:
            # último intento: buscar en las celdas de las primeras filas
            df_raw = _leer(None)
            for i in range(min(15, len(df_raw))):
                fila_vals = {str(v).strip().upper() for v in df_raw.iloc[i].values
                             if pd.notna(v)}
                if fila_vals & COLS_CLAVE:
                    df = _leer(i)
                    break

    # Normalizar nombres de columnas
    col_map = {}
    for c in df.columns:
        cl = str(c).strip().upper()
        if cl in ("NOMBRE", "ALUMNO", "APELLIDOS", "APELLIDOS, NOMBRE"):
            col_map[c] = "NOMBRE"
        elif cl in ("NACIONALIDAD", "NACION", "NACIONALITAT", "PAÍS"):
            col_map[c] = "NACIONALIDAD"
        elif cl in ("SEXO", "GÉNERO", "GENERO", "GENDER", "SEXE"):
            col_map[c] = "SEXO"
        elif cl in ("NACIMIENTO", "F.NAC", "FECHA_NAC", "FECHA NAC",
                    "FECHA DE NACIMIENTO", "F_NAC", "BIRTHDATE", "DOB"):
            col_map[c] = "NACIMIENTO"
    df = df.rename(columns=col_map)

    req = {"SEXO", "NACIMIENTO"}
    missing = req - set(df.columns)
    if missing:
        raise ValueError(f"No se encontraron las columnas: {', '.join(missing)}\n"
                         f"Columnas disponibles: {', '.join(df.columns)}")

    df["_FECHA"] = df["NACIMIENTO"].apply(_parse_date)
    ref = fecha_ref
    df["_EDAD"] = df["_FECHA"].apply(
        lambda d: ref.year - d.year - ((ref.month, ref.day) < (d.month, d.day))
        if d else None
    )
    df["_GRUPO"] = df["_EDAD"].apply(_age_bucket)

    sexo_col = df["SEXO"].str.strip().str.upper()
    df["_H"] = sexo_col.isin(["H", "M", "HOMBRE", "MALE", "1"])  # H o M (masculino)
    # Detectar convención H/M vs H/F
    vals = sexo_col.unique()
    if any(v in vals for v in ["M", "MUJER", "FEMALE", "F"]):
        if "M" in vals and "H" not in vals:
            # M=masculino, F=femenino
            df["_H"] = sexo_col.isin(["M", "MALE", "HOMBRE", "1"])
        elif "H" in vals:
            df["_H"] = sexo_col.isin(["H", "HOMBRE", "MALE", "1"])
    masc_vals = {"H", "HOMBRE", "MALE", "1", "M"} & set(vals)
    fem_vals  = {"M", "MUJER", "FEMALE", "F", "0", "2"} & set(vals)
    # Si hay ambas H y M, H=hombre M=mujer
    if "H" in vals and "M" in vals:
        df["_H"] = sexo_col == "H"
    elif "H" in vals:
        df["_H"] = sexo_col == "H"
    elif "M" in vals and "F" not in vals:
        df["_H"] = sexo_col == "M"
    else:
        df["_H"] = sexo_col.isin(["H", "HOMBRE", "MALE", "1"])

    total = len(df)
    hombres = df["_H"].sum()
    mujeres = total - hombres

    # Edades
    age_stats = {}
    all_buckets = [b for b in AGE_LABELS if b in df["_GRUPO"].values]
    for g in AGE_LABELS:
        sub = df[df["_GRUPO"] == g]
        if len(sub) > 0:
            age_stats[g] = {
                "total": len(sub),
                "h": int(sub["_H"].sum()),
                "m": int(len(sub) - sub["_H"].sum()),
            }

    # Nacionalidades
    nat_stats = {}
    if "NACIONALIDAD" in df.columns:
        df["_NAC"] = df["NACIONALIDAD"].str.strip().map(
            lambda x: NAT_MAP.get(x, x) if pd.notna(x) else "Desconocida"
        )
        nat_espanola = df[df["NACIONALIDAD"].str.strip().str.lower().str.contains(
            "españ|espany|spain|spanish", na=False, regex=True
        )]
        nat_extran = df[~df["NACIONALIDAD"].str.strip().str.lower().str.contains(
            "españ|espany|spain|spanish", na=False, regex=True
        )]
        for nat, grp in nat_extran.groupby("_NAC"):
            nat_stats[nat] = {
                "total": len(grp),
                "h": int(grp["_H"].sum()),
                "m": int(len(grp) - grp["_H"].sum()),
            }
        esp_t = len(nat_espanola)
        esp_h = int(nat_espanola["_H"].sum())
        nat_espanola_stats = {"total": esp_t, "h": esp_h, "m": esp_t - esp_h}
    else:
        nat_espanola_stats = None
        nat_extran = df

    return {
        "total": total,
        "hombres": int(hombres),
        "mujeres": int(mujeres),
        "age_stats": age_stats,
        "nat_stats": nat_stats,
        "nat_espanola": nat_espanola_stats,
        "archivo": os.path.basename(path),
        "fecha_ref": fecha_ref.strftime("%d/%m/%Y"),
        "fecha_gen": datetime.now().strftime("%d/%m/%Y"),
    }

# ── Generación XLSX ──────────────────────────────────────────────────────────

def generar_xlsx(datos, dest_path):
    # Estilos
    AZ      = "2563EB"
    AZ_CLARO= "DBEAFE"
    GRIS    = "F1F5F9"
    BLANCO  = "FFFFFF"
    OSCURO  = "1E293B"
    VERDE   = "16A34A"

    hdr_font  = Font(name="Calibri", bold=True, color=BLANCO, size=11)
    hdr_fill  = PatternFill("solid", fgColor=AZ)
    tot_font  = Font(name="Calibri", bold=True, color=OSCURO, size=11)
    tot_fill  = PatternFill("solid", fgColor=AZ_CLARO)
    dat_font  = Font(name="Calibri", size=11, color=OSCURO)
    alt_fill  = PatternFill("solid", fgColor=GRIS)
    bla_fill  = PatternFill("solid", fgColor=BLANCO)
    tit_font  = Font(name="Calibri", bold=True, size=13, color=AZ)
    sub_font  = Font(name="Calibri", size=10, color="64748B", italic=True)

    thin = Side(style="thin", color="CBD5E1")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    centro = Alignment(horizontal="center", vertical="center")
    izq    = Alignment(horizontal="left",   vertical="center", indent=1)

    wb = openpyxl.Workbook()

    def _hoja_tabla(nombre_hoja, titulo, subtitulo, headers, filas, anchos):
        ws = wb.create_sheet(nombre_hoja)
        ws.sheet_view.showGridLines = False

        # Título y subtítulo
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(headers))
        ws.cell(1, 1).value = titulo
        ws.cell(1, 1).font  = tit_font
        ws.cell(1, 1).alignment = izq
        ws.row_dimensions[1].height = 22

        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=len(headers))
        ws.cell(2, 1).value = subtitulo
        ws.cell(2, 1).font  = sub_font
        ws.cell(2, 1).alignment = izq
        ws.row_dimensions[2].height = 16

        # Cabecera
        for c, hdr in enumerate(headers, 1):
            cell = ws.cell(3, c, hdr)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.border    = borde
            cell.alignment = centro
        ws.row_dimensions[3].height = 18

        # Filas de datos
        for r, fila in enumerate(filas, 4):
            es_total = str(fila[0]).upper() == "TOTAL"
            for c, val in enumerate(fila, 1):
                cell = ws.cell(r, c, val)
                cell.border = borde
                cell.font   = tot_font if es_total else dat_font
                cell.fill   = tot_fill if es_total else (
                              alt_fill if (r % 2 == 0) else bla_fill)
                cell.alignment = izq if c == 1 else centro
            ws.row_dimensions[r].height = 17

        # Anchos de columna
        for c, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(c)].width = ancho

        return ws

    age_stats = datos["age_stats"]
    filas_edad = []
    for lbl in AGE_LABELS:
        v = age_stats.get(lbl, {"total": 0, "h": 0, "m": 0})
        filas_edad.append((lbl, v["total"], v["h"], v["m"]))
    filas_edad.append(("TOTAL", datos["total"], datos["hombres"], datos["mujeres"]))

    subtitulo_comun = (f"Fecha de referencia: {datos['fecha_ref']}   ·   "
                       f"Total: {datos['total']}   ·   "
                       f"Hombres: {datos['hombres']}   ·   "
                       f"Mujeres: {datos['mujeres']}   ·   "
                       f"Generado: {datos['fecha_gen']}")

    _hoja_tabla(
        "Por edad",
        "Alumnado matriculado por edad",
        subtitulo_comun,
        ["Tramo de edad", "Total", "Hombres", "Mujeres"],
        filas_edad,
        [22, 10, 10, 10],
    )

    nat_stats = datos["nat_stats"]
    nat_esp   = datos["nat_espanola"]
    if nat_stats or (nat_esp and nat_esp["total"] > 0):
        filas_nat = []
        if nat_esp and nat_esp["total"] > 0:
            filas_nat.append(("Española", nat_esp["total"], nat_esp["h"], nat_esp["m"]))
        for nat, vv in sorted(nat_stats.items(), key=lambda x: -x[1]["total"]):
            filas_nat.append((nat, vv["total"], vv["h"], vv["m"]))
        filas_nat.append(("TOTAL", datos["total"], datos["hombres"], datos["mujeres"]))

        _hoja_tabla(
            "Por nacionalidad",
            "Alumnado matriculado por nacionalidad",
            subtitulo_comun,
            ["Nacionalidad", "Total", "Hombres", "Mujeres"],
            filas_nat,
            [22, 10, 10, 10],
        )

    # Eliminar hoja vacía por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(dest_path)

# ── GUI ──────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Estadísticas de Alumnado → Excel")
        self.configure(bg=C_BG)
        self.resizable(False, False)

        # Centrar en pantalla
        self.update_idletasks()
        W_win, H_win = 560, 430
        x = (self.winfo_screenwidth()  - W_win) // 2
        y = max(40, (self.winfo_screenheight() - H_win) // 2 - 60)
        self.geometry(f"{W_win}x{H_win}+{x}+{y}")

        self._src_path  = tk.StringVar()
        self._dest_dir  = tk.StringVar(value=os.path.expanduser("~\\Desktop")
                                       if os.name == "nt" else os.path.expanduser("~/Escritorio"))
        self._fecha_ref = tk.StringVar(value=f"01/09/{datetime.now().year}")
        self._auto_open = tk.BooleanVar(value=True)
        self._status    = tk.StringVar(value="")

        self._build_ui()

    # ── Construcción UI ──────────────────────────────────────────────────────

    def _lbl(self, parent, text, size=10, bold=False, color=C_TEXT):
        return tk.Label(parent, text=text,
                        font=("Segoe UI", size, "bold" if bold else "normal"),
                        fg=color, bg=C_BG)

    def _build_ui(self):
        pad = dict(padx=24)

        # Título
        tk.Label(self, text="Estadísticas de Alumnado",
                 font=("Segoe UI", 15, "bold"), fg=C_ACCENT, bg=C_BG
                 ).pack(pady=(20, 2))
        tk.Label(self, text="Genera un Excel con estadísticas desde un listado de alumnos",
                 font=("Segoe UI", 9), fg=C_MUTED, bg=C_BG
                 ).pack(pady=(0, 18))

        # Separador
        ttk.Separator(self, orient="horizontal").pack(fill="x", **pad)

        # Sección: archivo fuente
        frm = tk.Frame(self, bg=C_BG)
        frm.pack(fill="x", pady=(16, 6), **pad)
        self._lbl(frm, "Archivo de alumnado  (.xls · .xlsx · .csv)", bold=True).grid(
            row=0, column=0, columnspan=3, sticky="w")
        self._entry(frm, self._src_path, row=1, browse_cmd=self._browse_src)

        # Sección: carpeta destino
        frm2 = tk.Frame(self, bg=C_BG)
        frm2.pack(fill="x", pady=(10, 6), **pad)
        self._lbl(frm2, "Carpeta de destino", bold=True).grid(
            row=0, column=0, columnspan=3, sticky="w")
        self._entry(frm2, self._dest_dir, row=1, browse_cmd=self._browse_dest,
                    folder=True)

        # Sección: fecha de referencia
        frm_f = tk.Frame(self, bg=C_BG)
        frm_f.pack(fill="x", pady=(10, 6), **pad)
        self._lbl(frm_f, "Fecha de referencia para el cálculo de edades", bold=True).grid(
            row=0, column=0, columnspan=3, sticky="w")
        frm_fi = tk.Frame(frm_f, bg=C_BG)
        frm_fi.grid(row=1, column=0, columnspan=3, sticky="w", pady=3)
        self._fecha_entry = tk.Entry(frm_fi, textvariable=self._fecha_ref,
                                     font=("Segoe UI", 9), fg=C_TEXT,
                                     relief="solid", bd=1, bg=C_PANEL, width=14,
                                     justify="center")
        self._fecha_entry.pack(side="left", ipady=4)
        tk.Label(frm_fi, text="  (DD/MM/AAAA  ·  normalmente 01/09 del curso)",
                 font=("Segoe UI", 9), fg=C_MUTED, bg=C_BG).pack(side="left")

        # Opción abrir
        frm3 = tk.Frame(self, bg=C_BG)
        frm3.pack(fill="x", pady=(8, 0), **pad)
        tk.Checkbutton(frm3, text="Abrir el Excel al terminar",
                       variable=self._auto_open,
                       font=("Segoe UI", 10), fg=C_TEXT, bg=C_BG,
                       activebackground=C_BG, selectcolor=C_BG
                       ).pack(side="left")

        # Botón generar
        self._btn = tk.Button(self, text="⚡  Generar Excel",
                              font=("Segoe UI", 11, "bold"),
                              bg=C_ACCENT, fg=C_BTN_FG, relief="flat",
                              cursor="hand2", padx=20, pady=8,
                              command=self._run)
        self._btn.pack(pady=(18, 6))

        # Estado
        self._status_lbl = tk.Label(self, textvariable=self._status,
                                    font=("Segoe UI", 9), fg=C_MUTED, bg=C_BG,
                                    wraplength=500)
        self._status_lbl.pack(pady=(0, 12))

    def _entry(self, parent, var, row, browse_cmd, folder=False):
        e = tk.Entry(parent, textvariable=var,
                     font=("Segoe UI", 9), fg=C_TEXT,
                     relief="solid", bd=1, bg=C_PANEL, width=46)
        e.grid(row=row, column=0, sticky="ew", ipady=4, pady=3)
        btn = tk.Button(parent, text="…",
                        font=("Segoe UI", 9), bg=C_BORDER, fg=C_TEXT,
                        relief="flat", cursor="hand2", padx=8,
                        command=browse_cmd)
        btn.grid(row=row, column=1, padx=(6, 0))
        parent.columnconfigure(0, weight=1)

    # ── Callbacks ────────────────────────────────────────────────────────────

    def _browse_src(self):
        path = filedialog.askopenfilename(
            title="Seleccionar listado de alumnos",
            filetypes=[("Excel / CSV", "*.xls *.xlsx *.xlsm *.csv"),
                       ("Todos", "*.*")])
        if path:
            self._src_path.set(path)
            # Proponer mismo directorio como destino si estaba vacío
            if not self._dest_dir.get():
                self._dest_dir.set(os.path.dirname(path))

    def _browse_dest(self):
        path = filedialog.askdirectory(title="Seleccionar carpeta de destino")
        if path:
            self._dest_dir.set(path)

    def _run(self):
        src  = self._src_path.get().strip()
        dest = self._dest_dir.get().strip()
        fecha_str = self._fecha_ref.get().strip()

        if not src:
            messagebox.showwarning("Faltan datos", "Selecciona el archivo de alumnado.")
            return
        if not os.path.isfile(src):
            messagebox.showerror("Error", f"No se encuentra el archivo:\n{src}")
            return
        if not dest:
            messagebox.showwarning("Faltan datos", "Selecciona la carpeta de destino.")
            return
        # Validar fecha
        try:
            ref = datetime.strptime(fecha_str, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Fecha incorrecta",
                                 f"El formato de la fecha debe ser DD/MM/AAAA.\n"
                                 f"Valor introducido: «{fecha_str}»")
            self._fecha_entry.focus_set()
            return
        if not os.path.isdir(dest):
            try:
                os.makedirs(dest)
            except Exception as e:
                messagebox.showerror("Error", f"No se puede crear la carpeta:\n{e}")
                return

        self._btn.config(state="disabled", text="Procesando…")
        self._set_status("Leyendo listado de alumnos…", C_MUTED)
        threading.Thread(target=self._worker, args=(src, dest, ref), daemon=True).start()

    def _worker(self, src, dest, ref):
        try:
            datos = procesar_listado(src, ref)
            self.after(0, self._set_status, "Generando Excel…", C_MUTED)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre = f"Estadisticas_{ts}.xlsx"
            out = os.path.join(dest, nombre)
            generar_xlsx(datos, out)
            msg = (f"✔  Excel guardado: {nombre}\n"
                   f"   {datos['total']} alumnos · {datos['hombres']}H · {datos['mujeres']}M")
            self.after(0, self._set_status, msg, C_SUCCESS)
            self.after(0, self._btn.config, {"state": "normal", "text": "⚡  Generar Excel"})
            if self._auto_open.get():
                self.after(500, self._abrir, out)
        except Exception as e:
            import traceback
            try:
                log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "estadisticas_error.log")
                with open(log_path, "w", encoding="utf-8") as f:
                    f.write(traceback.format_exc())
            except Exception:
                pass
            msg = f"✘  Error: {e}\n(Ver estadisticas_error.log para detalles)"
            self.after(0, self._set_status, msg, "#DC2626")
            self.after(0, self._btn.config, {"state": "normal", "text": "⚡  Generar Excel"})

    def _set_status(self, msg, color):
        self._status.set(msg)
        self._status_lbl.config(fg=color)

    def _abrir(self, path):
        try:
            if os.name == "nt":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception:
            pass


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
